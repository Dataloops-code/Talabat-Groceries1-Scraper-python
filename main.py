import asyncio
import json
import os
import tempfile
import sys
import subprocess
import re
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from SavingOnDrive import SavingOnDrive
import logging
from datetime import datetime
from retry import retry

# Set up logging (used by MainScraper)
logging.basicConfig(
    filename='scraper.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class TalabatGroceries:
    def __init__(self, url, browser):
        self.url = url
        self.base_url = "https://www.talabat.com"
        self.browser = browser
        print(f"Initialized TalabatGroceries with URL: {self.url}")

    async def get_general_link(self, page):
        print("Attempting to get general link")
        retries = 3
        while retries > 0:
            try:
                link_element = await page.wait_for_selector('//a[@data-testid="view-all-link"]', timeout=60000)
                if link_element:
                    full_link = self.base_url + await link_element.get_attribute('href')
                    print(f"General link found: {full_link}")
                    return full_link
                else:
                    print("General link not found")
                    return None
            except Exception as e:
                print(f"Error getting general link: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return None

    async def get_delivery_fees(self, page):
        print("Attempting to get delivery fees")
        retries = 3
        while retries > 0:
            try:
                delivery_fees_element = await page.query_selector('xpath=/html/body/div/div/div[1]/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[2]/span[1]')
                delivery_fees = await delivery_fees_element.inner_text() if delivery_fees_element else "N/A"
                print(f"Delivery fees: {delivery_fees}")
                return delivery_fees
            except Exception as e:
                print(f"Error getting delivery fees: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return "N/A"

    async def get_minimum_order(self, page):
        print("Attempting to get minimum order")
        retries = 3
        while retries > 0:
            try:
                minimum_order_element = await page.query_selector('xpath=/html/body/div/div/div[1]/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[2]/span[3]')
                minimum_order = await minimum_order_element.inner_text() if minimum_order_element else "N/A"
                print(f"Minimum order: {minimum_order}")
                return minimum_order
            except Exception as e:
                print(f"Error getting minimum order: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return "N/A"

    async def extract_category_names(self, page):
        print("Attempting to extract category names")
        retries = 3
        while retries > 0:
            try:
                category_name_elements = await page.query_selector_all('//span[@data-testid="category-name"]')
                category_names = [await element.inner_text() for element in category_name_elements]
                print(f"Category names extracted: {category_names}")
                return category_names
            except Exception as e:
                print(f"Error extracting category names: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return []

    async def extract_category_links(self, page):
        print("Attempting to extract category links")
        retries = 3
        while retries > 0:
            try:
                category_link_elements = await page.query_selector_all('//a[@data-testid="category-item-container"]')
                category_links = [self.base_url + await element.get_attribute('href') for element in category_link_elements]
                print(f"Category links extracted: {category_links}")
                return category_links
            except Exception as e:
                print(f"Error extracting category links: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return []

    async def extract_sub_categories(self, page, category_link, main_scraper, grocery_title, category_info, last_sub_category=None):
        """
        Extract sub-categories, starting at the next sub-category after last_sub_category.
        Updates current_sub_category to the next unprocessed sub-category after completion.
        """
        print(f"Attempting to extract sub-categories for: {category_link}")
        retries = 3
        sub_categories = []
        completed_sub_categories = main_scraper.current_progress["current_progress"]["completed_categories"].get(grocery_title, {}).get(category_info["name"], [])
        is_resuming = last_sub_category is not None

        while retries > 0:
            try:
                await page.goto(category_link, timeout=240000)
                await page.wait_for_load_state("networkidle", timeout=240000)
                sub_category_elements = await page.query_selector_all('//div[@data-test="sub-category-container"]//a[@data-testid="subCategory-a"]')
                sub_category_names = [await el.inner_text() for el in sub_category_elements]
                sub_category_links = [self.base_url + await el.get_attribute('href') for el in sub_category_elements]

                # Find the next sub-category to process
                start_processing = not is_resuming
                for idx, (sub_category_name, sub_category_link) in enumerate(zip(sub_category_names, sub_category_links)):
                    # Skip completed sub-categories
                    if sub_category_name in completed_sub_categories:
                        print(f"    Skipping completed sub-category: {sub_category_name}")
                        continue

                    # If resuming, skip until after last_sub_category
                    if is_resuming and not start_processing:
                        if sub_category_name == last_sub_category:
                            print(f"    Found last processed sub-category: {sub_category_name}, starting next")
                            start_processing = True
                        continue

                    if not start_processing:
                        continue

                    print(f"    Processing sub-category: {sub_category_name}")
                    print(f"    Sub-category link: {sub_category_link}")
                    items = await self.extract_all_items_from_sub_category(sub_category_link)
                    sub_category_data = {
                        "sub_category_name": sub_category_name,
                        "sub_category_link": sub_category_link,
                        "Items": items
                    }
                    sub_categories.append(sub_category_data)

                    # Update progress to mark this sub-category complete
                    completed_cats = main_scraper.current_progress["current_progress"]["completed_categories"].setdefault(grocery_title, {})
                    completed_cats.setdefault(category_info["name"], []).append(sub_category_name)
                    main_scraper.scraped_progress["current_progress"]["completed_categories"] = main_scraper.current_progress["current_progress"]["completed_categories"]

                    # Set current_sub_category to the NEXT sub-category (or None if last)
                    next_sub_idx = idx + 1
                    if next_sub_idx < len(sub_category_names):
                        next_sub_name = sub_category_names[next_sub_idx]
                        if next_sub_name not in completed_sub_categories:
                            main_scraper.current_progress["current_progress"]["current_sub_category"] = next_sub_name
                            main_scraper.scraped_progress["current_progress"]["current_sub_category"] = next_sub_name
                        else:
                            main_scraper.current_progress["current_progress"]["current_sub_category"] = None
                            main_scraper.scraped_progress["current_progress"]["current_sub_category"] = None
                    else:
                        main_scraper.current_progress["current_progress"]["current_sub_category"] = None
                        main_scraper.scraped_progress["current_progress"]["current_sub_category"] = None

                    # Update scraped_progress results
                    area_name = main_scraper.current_progress["current_progress"]["area_name"]
                    if area_name:
                        grocery_data = main_scraper.scraped_progress["all_results"].setdefault(area_name, {}).setdefault(grocery_title, {
                            "grocery_link": self.url,
                            "delivery_time": "N/A",
                            "grocery_details": {"delivery_fees": "N/A", "minimum_order": "N/A", "categories": []}
                        })
                        if not grocery_data["grocery_details"]["categories"]:
                            grocery_data["grocery_details"]["categories"] = [{"name": cat["name"], "link": cat["link"], "sub_categories": []} for cat in category_info.get("categories", [category_info])]
                        for cat in grocery_data["grocery_details"]["categories"]:
                            if cat["name"] == category_info["name"]:
                                cat["sub_categories"] = sub_categories
                                break
                        main_scraper.scraped_progress["all_results"][area_name][grocery_title] = grocery_data

                    # Save and commit progress
                    main_scraper.save_current_progress()
                    main_scraper.save_scraped_progress()
                    main_scraper.commit_progress(f"Processed sub-category {sub_category_name} for {grocery_title} in {category_info['name']}")

                print(f"Extracted {len(sub_categories)} new sub-categories")
                return sub_categories
            except Exception as e:
                print(f"Error extracting sub-categories: {e}")
                logging.error(f"Error extracting sub-categories: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return sub_categories
        
    async def extract_item_details(self, item_link):
        print(f"Attempting to extract item details for link: {item_link}")
        retries = 3
        while retries > 0:
            try:
                page = await self.browser.new_page()
                await page.goto(item_link, timeout=240000)
                await page.wait_for_load_state("networkidle", timeout=240000)
                await page.wait_for_timeout(5000)
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_load_state("networkidle", timeout=240000)

                price_selectors = [
                    '//div[@class="price"]//span[@class="currency "]',
                    '//span[contains(@class, "price")]',
                    '//div[contains(@class, "price")]//text()'
                ]
                item_price = "N/A"
                for selector in price_selectors:
                    price_element = await page.query_selector(selector)
                    if price_element:
                        item_price = await price_element.inner_text()
                        break

                desc_selectors = [
                    '//div[@class="description"]//p[@data-testid="item-description"]',
                    '//div[contains(@class, "description")]//p',
                    '//p[contains(@class, "description")]'
                ]
                item_description = "N/A"
                for selector in desc_selectors:
                    desc_element = await page.query_selector(selector)
                    if desc_element:
                        item_description = await desc_element.inner_text()
                        break

                delivery_time_element = await page.query_selector('//div[@data-testid="delivery-tag"]//span')
                delivery_time = await delivery_time_element.inner_text() if delivery_time_element else "N/A"

                image_selectors = [
                    '//div[@data-testid="item-image"]//img',
                    '//img[contains(@class, "item-image")]',
                    '//img[@alt="product image"]'
                ]
                item_images = []
                for selector in image_selectors:
                    item_image_elements = await page.query_selector_all(selector)
                    if item_image_elements:
                        item_images = [await img.get_attribute('src') for img in item_image_elements if await img.get_attribute('src')]
                        break

                print(f"Item price: {item_price}")
                print(f"Item description: {item_description}")
                print(f"Delivery time range: {delivery_time}")
                print(f"Item images: {item_images}")

                await page.close()
                return {
                    "item_price": item_price,
                    "item_description": item_description,
                    "item_delivery_time_range": delivery_time,
                    "item_images": item_images
                }
            except Exception as e:
                print(f"Error extracting item details for {item_link}: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        print(f"Failed to extract details for {item_link} after all retries")
        return {
            "item_price": "N/A",
            "item_description": "N/A",
            "item_delivery_time_range": "N/A",
            "item_images": []
        }

    async def extract_all_items_from_sub_category(self, sub_category_link):
        print(f"Attempting to extract all items from sub-category: {sub_category_link}")
        retries = 3
        while retries > 0:
            try:
                sub_page = await self.browser.new_page()
                await sub_page.goto(sub_category_link, timeout=240000)
                await sub_page.wait_for_load_state("networkidle", timeout=240000)
                await sub_page.wait_for_selector('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]', timeout=240000)

                pagination_element = await sub_page.query_selector('//div[@class="sc-104fa483-0 fCcIDQ"]//ul[@class="paginate-wrap"]')
                total_pages = 1
                if pagination_element:
                    page_numbers = await pagination_element.query_selector_all('//li[contains(@class, "paginate-li f-16 f-500")]//a')
                    total_pages = len(page_numbers) if page_numbers else 1
                print(f"      Found {total_pages} pages in this sub-category")

                items = []
                for page_number in range(1, total_pages + 1):
                    print(f"      Processing page {page_number} of {total_pages}")
                    page_url = f"{sub_category_link}&page={page_number}"
                    await sub_page.goto(page_url, timeout=240000)
                    await sub_page.wait_for_load_state("networkidle", timeout=240000)
                    await sub_page.wait_for_selector('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]', timeout=240000)

                    item_elements = await sub_page.query_selector_all('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]//a[@data-testid="grocery-item-link-nofollow"]')
                    print(f"        Found {len(item_elements)} items on page {page_number}")

                    for i, element in enumerate(item_elements):
                        try:
                            item_name_element = await element.query_selector('div[data-test="item-name"]')
                            item_name = await item_name_element.inner_text() if item_name_element else f"Unknown Item {i+1}"
                            print(f"        Item name: {item_name}")

                            item_link = self.base_url + await element.get_attribute('href')
                            print(f"        Item link: {item_link}")

                            item_details = await self.extract_item_details(item_link)
                            items.append({
                                "item_name": item_name,
                                "item_link": item_link,
                                **item_details
                            })
                        except Exception as e:
                            print(f"        Error processing item {i+1}: {e}")
                await sub_page.close()
                return items
            except Exception as e:
                print(f"Error extracting items from sub-category {sub_category_link}: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return []

    async def extract_categories(self, page):
        print(f"Processing grocery: {self.url}")
        retries = 3
        while retries > 0:
            try:
                await page.goto(self.url, timeout=240000)
                await page.wait_for_load_state("networkidle", timeout=240000)
                print("Page loaded successfully")

                delivery_fees = await self.get_delivery_fees(page)
                minimum_order = await self.get_minimum_order(page)
                view_all_link = await self.get_general_link(page)

                print(f"  Delivery fees: {delivery_fees}")
                print(f"  Minimum order: {minimum_order}")

                categories_data = []
                if view_all_link:
                    print(f"  Navigating to view all link: {view_all_link}")
                    category_page = await self.browser.new_page()
                    await category_page.goto(view_all_link, timeout=240000)
                    await category_page.wait_for_load_state("networkidle", timeout=240000)

                    category_names = await self.extract_category_names(category_page)
                    category_links = await self.extract_category_links(category_page)

                    print(f"  Found {len(category_names)} categories")

                    for index, (name, link) in enumerate(zip(category_names, category_links)):
                        print(f"  Processing category {index+1}/{len(category_names)}: {name}")
                        print(f"  Category link: {link}")
                        categories_data.append({
                            "name": name,
                            "link": link,
                            "sub_categories": []
                        })
                    await category_page.close()

                return {
                    "delivery_fees": delivery_fees,
                    "minimum_order": minimum_order,
                    "categories": categories_data
                }
            except Exception as e:
                print(f"Error extracting categories: {e}")
                retries -= 1
                print(f"Retries left: {retries}")
                await asyncio.sleep(5)
        return {"error": "Failed to extract categories after multiple attempts"}



class MainScraper:
    CURRENT_PROGRESS_FILE = "current_progress.json"
    SCRAPED_PROGRESS_FILE = "scraped_progress.json"

    def __init__(self):
        self.output_dir = "output"
        credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
        # Validate credentials_json
        if not credentials_json:
            logging.warning("TALABAT_GCLOUD_KEY_JSON is not set. Google Drive uploads will fail.")
            self.drive_uploader = SavingOnDrive(credentials_json=None)
        else:
            try:
                credentials_dict = json.loads(credentials_json)
                if not credentials_dict.get('type') == 'service_account':
                    logging.warning("TALABAT_GCLOUD_KEY_JSON is not a valid service account key. Google Drive uploads will fail.")
                    self.drive_uploader = SavingOnDrive(credentials_json=None)
                else:
                    self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
            except json.JSONDecodeError as e:
                logging.warning(f"TALABAT_GCLOUD_KEY_JSON is invalid JSON: {str(e)}. Google Drive uploads will fail.")
                self.drive_uploader = SavingOnDrive(credentials_json=None)
        
        os.makedirs(self.output_dir, exist_ok=True)
        self.current_progress = self.load_current_progress()
        self.scraped_progress = self.load_scraped_progress()
        self.github_token = os.environ.get('GITHUB_TOKEN')
        if not self.github_token:
            logging.warning("GITHUB_TOKEN is not set. Git pushes will be skipped.")
        self.ensure_playwright_browsers()
        # Initialize progress files if empty, matching Restaurants
        self.commit_progress("Initialized progress files at scraper start")

    def ensure_playwright_browsers(self):
        subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)

    def load_current_progress(self) -> Dict:
        """
        Load current_progress.json, creating a default if missing, matching Restaurants logic.
        Deduplicates processed_groceries to prevent re-scraping.
        """
        if os.path.exists(self.CURRENT_PROGRESS_FILE):
            try:
                with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                # Validate and clean progress
                if "current_progress" not in progress:
                    progress["current_progress"] = {}
                current = progress["current_progress"]
                current.setdefault("processed_groceries", [])
                current.setdefault("completed_categories", {})
                current.setdefault("area_name", None)
                current.setdefault("current_grocery", 0)
                current.setdefault("current_grocery_title", None)
                current.setdefault("current_grocery_link", None)
                current.setdefault("current_category", None)
                current.setdefault("current_sub_category", None)
                current.setdefault("total_groceries", 0)
                # Deduplicate processed_groceries (by grocery_link)
                current["processed_groceries"] = list(set(current["processed_groceries"]))
                progress["last_updated"] = progress.get("last_updated", datetime.now().isoformat())
                logging.info(f"Loaded current_progress.json: {json.dumps(progress, indent=2)}")
                return progress
            except Exception as e:
                logging.error(f"Error loading current_progress.json: {e}")
        
        # Create default progress, matching Restaurants
        default_progress = {
            "completed_areas": [],
            "current_area_index": 0,
            "last_updated": datetime.now().isoformat(),
            "current_progress": {
                "area_name": None,
                "current_grocery": 0,
                "current_grocery_title": None,
                "current_grocery_link": None,
                "total_groceries": 0,
                "processed_groceries": [],
                "current_category": None,
                "current_sub_category": None,
                "completed_categories": {}
            }
        }
        logging.info("current_progress.json not found, creating default")
        self.save_current_progress(default_progress)
        self.commit_progress("Created default current_progress.json")
        return default_progress

    def save_current_progress(self, progress: Dict = None):
        """
        Save current_progress.json with atomic write, matching Restaurants.
        Updates last_updated and deduplicates processed_groceries.
        """
        progress = progress or self.current_progress
        try:
            progress["last_updated"] = datetime.now().isoformat()
            if "current_progress" in progress:
                progress["current_progress"]["processed_groceries"] = list(set(progress["current_progress"].get("processed_groceries", [])))
            with tempfile.NamedTemporaryFile('w', delete=False, dir='.', encoding='utf-8') as temp_file:
                json.dump(progress, temp_file, indent=2, ensure_ascii=False)
                temp_file.flush()
                os.fsync(temp_file.fileno())
                temp_filename = temp_file.name
            os.replace(temp_filename, self.CURRENT_PROGRESS_FILE)
            logging.info(f"Saved current_progress.json: {self.CURRENT_PROGRESS_FILE}")
        except Exception as e:
            logging.error(f"Error saving current_progress.json: {e}")

    def load_scraped_progress(self) -> Dict:
        """
        Load scraped_progress.json, creating a default if missing, matching Restaurants.
        Initializes all_results and deduplicates processed_groceries.
        """
        if os.path.exists(self.SCRAPED_PROGRESS_FILE):
            try:
                with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                # Validate and clean progress
                if "current_progress" not in progress:
                    progress["current_progress"] = {}
                if "all_results" not in progress:
                    progress["all_results"] = {}
                current = progress["current_progress"]
                current.setdefault("processed_groceries", [])
                current.setdefault("completed_categories", {})
                current.setdefault("area_name", None)
                current.setdefault("current_grocery", 0)
                current.setdefault("current_grocery_title", None)
                current.setdefault("current_grocery_link", None)
                current.setdefault("current_category", None)
                current.setdefault("current_sub_category", None)
                current.setdefault("total_groceries", 0)
                # Deduplicate processed_groceries
                current["processed_groceries"] = list(set(current["processed_groceries"]))
                progress["last_updated"] = progress.get("last_updated", datetime.now().isoformat())
                logging.info(f"Loaded scraped_progress.json: {json.dumps(progress, indent=2)}")
                return progress
            except Exception as e:
                logging.error(f"Error loading scraped_progress.json: {e}")
        
        # Create default progress, matching Restaurants
        default_progress = {
            "completed_areas": [],
            "current_area_index": 0,
            "last_updated": datetime.now().isoformat(),
            "all_results": {},
            "current_progress": {
                "area_name": None,
                "current_grocery": 0,
                "current_grocery_title": None,
                "current_grocery_link": None,
                "total_groceries": 0,
                "processed_groceries": [],
                "current_category": None,
                "current_sub_category": None,
                "completed_categories": {}
            }
        }
        logging.info("scraped_progress.json not found, creating default")
        self.save_scraped_progress(default_progress)
        self.commit_progress("Created default scraped_progress.json")
        return default_progress

    def save_scraped_progress(self, progress: Dict = None):
        """
        Save scraped_progress.json with atomic write, matching Restaurants.
        Updates last_updated and deduplicates processed_groceries.
        """
        progress = progress or self.scraped_progress
        try:
            progress["last_updated"] = datetime.now().isoformat()
            if "current_progress" in progress:
                progress["current_progress"]["processed_groceries"] = list(set(progress["current_progress"].get("processed_groceries", [])))
            with tempfile.NamedTemporaryFile('w', delete=False, dir='.', encoding='utf-8') as temp_file:
                json.dump(progress, temp_file, indent=2, ensure_ascii=False)
                temp_file.flush()
                os.fsync(temp_file.fileno())
                temp_filename = temp_file.name
            os.replace(temp_filename, self.SCRAPED_PROGRESS_FILE)
            logging.info(f"Saved scraped_progress.json: {self.SCRAPED_PROGRESS_FILE}")
        except Exception as e:
            logging.error(f"Error saving scraped_progress.json: {e}")

    @retry(tries=3, delay=2, backoff=2)
    def commit_progress(self, message: str = "Periodic progress update"):
        """
        Commit progress files to Git, matching Restaurants logic.
        Retries on failure and logs outcomes.
        """
        if not self.github_token:
            logging.warning("No GITHUB_TOKEN, skipping commit")
            return
        try:
            logging.info(f"Attempting to commit progress: {message}")
            subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE, self.SCRAPED_PROGRESS_FILE, self.output_dir], check=True)
            result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
            if result.returncode == 0:
                subprocess.run(["git", "push"], check=True, env={"GIT_AUTH_TOKEN": self.github_token, **os.environ})
                logging.info(f"Successfully committed and pushed: {message}")
            else:
                logging.info(f"No changes to commit: {result.stdout}")
        except subprocess.CalledProcessError as e:
            logging.error(f"Error committing progress: {e}")
            raise

    async def process_grocery_categories(self, grocery_title, grocery_details, talabat_grocery, page, last_category=None, last_sub_category=None):
        """
        Process categories, starting at the next unprocessed category after last_category.
        Sets current_category to the next unprocessed category after completion.
        """
        completed_categories = self.current_progress["current_progress"]["completed_categories"].get(grocery_title, {})
        is_resuming = last_category is not None
        start_processing = not is_resuming

        categories = grocery_details.get("categories", [])
        for idx, category in enumerate(categories):
            category_name = category["name"]
            completed_sub_cats = completed_categories.get(category_name, [])

            # Skip if resuming until after last_category
            if is_resuming and not start_processing:
                if category_name == last_category:
                    print(f"Found last processed category: {category_name}, starting next")
                    start_processing = True
                continue

            # Skip fully completed categories
            if "sub_categories" in category and all(sub_cat["sub_category_name"] in completed_sub_cats for sub_cat in category["sub_categories"]):
                print(f"Skipping fully completed category: {category_name}")
                continue

            # Process category
            last_sub_cat_for_category = last_sub_category if category_name == last_category and is_resuming else None
            self.current_progress["current_progress"]["current_category"] = category_name
            self.scraped_progress["current_progress"]["current_category"] = category_name
            await self.process_category(grocery_title, category, talabat_grocery, page, last_sub_cat_for_category)

            # Set current_category to the next unprocessed category (or None if last)
            next_cat_idx = idx + 1
            if next_cat_idx < len(categories):
                next_cat_name = categories[next_cat_idx]["name"]
                next_sub_cats = categories[next_cat_idx].get("sub_categories", [])
                if not all(sub_cat["sub_category_name"] in completed_categories.get(next_cat_name, []) for sub_cat in next_sub_cats):
                    self.current_progress["current_progress"]["current_category"] = next_cat_name
                    self.scraped_progress["current_progress"]["current_category"] = next_cat_name
                else:
                    self.current_progress["current_progress"]["current_category"] = None
                    self.scraped_progress["current_progress"]["current_category"] = None
            else:
                self.current_progress["current_progress"]["current_category"] = None
                self.scraped_progress["current_progress"]["current_category"] = None

            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Completed category {category_name} for {grocery_title}")

    async def scrape_and_save_area(self, area_name: str, area_url: str, browser) -> List[Dict]:
        """
        Scrape groceries, starting at the next unprocessed grocery after last_grocery_link.
        Marks groceries complete only after all categories are processed.
        """
        print(f"\n{'='*50}\nSCRAPING AREA: {area_name}\nURL: {area_url}\n{'='*50}")
        all_area_results = self.scraped_progress["all_results"].get(area_name, {})
        current_progress = self.current_progress["current_progress"]
        scraped_current_progress = self.scraped_progress["current_progress"]

        is_resuming = current_progress["area_name"] == area_name
        last_grocery_link = current_progress.get("current_grocery_link")
        last_category = current_progress.get("current_category")
        last_sub_category = current_progress.get("current_sub_category")

        if not is_resuming:
            current_progress.update({
                "area_name": area_name,
                "current_grocery": 0,
                "current_grocery_title": None,
                "current_grocery_link": None,
                "total_groceries": 0,
                "processed_groceries": [],
                "current_category": None,
                "current_sub_category": None,
                "completed_categories": {}
            })
            scraped_current_progress.update(current_progress)
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Started scraping {area_name}")
        else:
            logging.info(f"Resuming area {area_name} from grocery link {last_grocery_link}, category {last_category}, sub-category {last_sub_category}")

        page = await browser.new_page()
        await page.goto(area_url, timeout=60000)
        groceries_on_page = await self.get_page_groceries(page)
        current_progress["total_groceries"] = len(groceries_on_page)
        scraped_current_progress["total_groceries"] = len(groceries_on_page)
        print(f"Found {len(groceries_on_page)} groceries")
        await page.close()

        processed_grocery_links = set(current_progress["processed_groceries"])
        start_processing = not is_resuming

        for grocery_idx, grocery in enumerate(groceries_on_page):
            grocery_num = grocery_idx + 1
            grocery_title = grocery["grocery_title"]
            grocery_link = grocery["grocery_link"]

            # Skip processed groceries
            if grocery_link in processed_grocery_links:
                logging.info(f"Skipping already processed grocery {grocery_title} (link: {grocery_link})")
                continue

            # Skip until after last_grocery_link
            if is_resuming and not start_processing:
                if grocery_link == last_grocery_link:
                    print(f"Found last processed grocery: {grocery_title}, starting next")
                    start_processing = True
                continue

            if not start_processing:
                continue

            current_progress["current_grocery"] = grocery_num
            current_progress["current_grocery_title"] = grocery_title
            current_progress["current_grocery_link"] = grocery_link
            scraped_current_progress["current_grocery"] = grocery_num
            scraped_current_progress["current_grocery_title"] = grocery_title
            scraped_current_progress["current_grocery_link"] = grocery_link
            print(f"Processing grocery {grocery_num}/{len(groceries_on_page)}: {grocery_title} (link: {grocery_link})")

            grocery_page = await browser.new_page()
            talabat_grocery = TalabatGroceries(grocery["grocery_link"], browser)
            grocery_details = await talabat_grocery.extract_categories(grocery_page)
            all_area_results[grocery_title] = {
                "grocery_link": grocery_link,
                "delivery_time": grocery["delivery_time"],
                "grocery_details": grocery_details
            }

            # Process categories
            last_cat = last_category if grocery_link == last_grocery_link and is_resuming else None
            last_sub_cat = last_sub_category if grocery_link == last_grocery_link and is_resuming else None
            await self.process_grocery_categories(grocery_title, grocery_details, talabat_grocery, grocery_page, last_cat, last_sub_cat)

            # Mark grocery as processed only if all categories are complete
            completed_cats = self.current_progress["current_progress"]["completed_categories"].get(grocery_title, {})
            is_grocery_complete = all(
                all(sub_cat["sub_category_name"] in completed_cats.get(cat["name"], [])
                    for sub_cat in cat.get("sub_categories", []))
                for cat in grocery_details.get("categories", [])
            )
            if is_grocery_complete:
                current_progress["processed_groceries"].append(grocery_link)
                scraped_current_progress["processed_groceries"].append(grocery_link)
                processed_grocery_links.add(grocery_link)
                # Set current_grocery to the next unprocessed grocery
                next_grocery_idx = grocery_idx + 1
                if next_grocery_idx < len(groceries_on_page):
                    next_grocery = groceries_on_page[next_grocery_idx]
                    if next_grocery["grocery_link"] not in processed_grocery_links:
                        current_progress["current_grocery"] = next_grocery_idx + 1
                        current_progress["current_grocery_title"] = next_grocery["grocery_title"]
                        current_progress["current_grocery_link"] = next_grocery["grocery_link"]
                        scraped_current_progress["current_grocery"] = next_grocery_idx + 1
                        scraped_current_progress["current_grocery_title"] = next_grocery["grocery_title"]
                        scraped_current_progress["current_grocery_link"] = next_grocery["grocery_link"]
                    else:
                        current_progress["current_grocery"] = 0
                        current_progress["current_grocery_title"] = None
                        current_progress["current_grocery_link"] = None
                        scraped_current_progress["current_grocery"] = 0
                        scraped_current_progress["current_grocery_title"] = None
                        scraped_current_progress["current_grocery_link"] = None
                else:
                    current_progress["current_grocery"] = 0
                    current_progress["current_grocery_title"] = None
                    current_progress["current_grocery_link"] = None
                    scraped_current_progress["current_grocery"] = 0
                    scraped_current_progress["current_grocery_title"] = None
                    scraped_current_progress["current_grocery_link"] = None

            self.scraped_progress["all_results"][area_name] = all_area_results
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Processed {grocery_title} (link: {grocery_link})")

            await grocery_page.close()

        json_filename = os.path.join(self.output_dir, f"{area_name}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(all_area_results, f, indent=2, ensure_ascii=False)

        workbook = Workbook()
        self.create_excel_sheet(workbook, area_name, all_area_results)
        excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
        workbook.save(excel_filename)
        if self.upload_to_drive(excel_filename):
            logging.info(f"Uploaded {excel_filename} to Google Drive")
        else:
            logging.warning(f"Failed to upload {excel_filename} to Google Drive")

        current_progress.update({
            "area_name": None,
            "current_grocery": 0,
            "current_grocery_title": None,
            "current_grocery_link": None,
            "total_groceries": 0,
            "processed_groceries": [],
            "current_category": None,
            "current_sub_category": None,
            "completed_categories": {}
        })
        scraped_current_progress.update(current_progress)
        self.save_current_progress()
        self.save_scraped_progress()
        self.commit_progress(f"Completed {area_name}")
        return list(all_area_results.values())
        
    async def process_category(self, grocery_title, category_info, talabat_grocery, page, last_sub_category=None):
        category_name = category_info["name"]
        self.current_progress["current_progress"]["current_category"] = category_name
        self.scraped_progress["current_progress"]["current_category"] = category_name
        sub_categories = await talabat_grocery.extract_sub_categories(page, category_info["link"], self, grocery_title, category_info, last_sub_category)
        category_info["sub_categories"] = sub_categories
        self.save_current_progress()
        self.save_scraped_progress()

    async def get_page_groceries(self, page) -> List[Dict]:
        logging.info("Extracting grocery information")
        try:
            await page.wait_for_selector('div[data-testid="one-vendor-container"]', timeout=30000)
            vendor_containers = await page.query_selector_all('div[data-testid="one-vendor-container"]')
            groceries_info = []
            for container in vendor_containers:
                title_element = await container.query_selector('a div h2')
                title = await title_element.inner_text() if title_element else "Unknown Grocery"
                link_element = await container.query_selector('a')
                link = "https://www.talabat.com" + await link_element.get_attribute('href') if link_element else None
                delivery_info = await container.query_selector('div.deliveryInfo')
                delivery_time_text = await delivery_info.inner_text() if delivery_info else ""
                delivery_time = re.findall(r'\d+', delivery_time_text)[0] + " mins" if re.findall(r'\d+', delivery_time_text) else "N/A"
                if link:
                    groceries_info.append({"grocery_title": title, "grocery_link": link, "delivery_time": delivery_time})
            logging.info(f"Extracted {len(groceries_info)} groceries: {[g['grocery_title'] for g in groceries_info]}")
            return groceries_info
        except Exception as e:
            logging.error(f"Error extracting groceries: {e}")
            return []

    def create_excel_sheet(self, workbook, sheet_name: str, data: Dict):
        sheet = workbook.create_sheet(title=sheet_name)
        simplified_data = []
        for grocery_title, grocery_data in data.items():
            general_info = {
                "Grocery Title": grocery_title,
                "Delivery Time": grocery_data.get("delivery_time", "N/A"),
                "Delivery Fees": grocery_data.get("grocery_details", {}).get("delivery_fees", "N/A"),
                "Minimum Order": grocery_data.get("grocery_details", {}).get("minimum_order", "N/A"),
                "URL": grocery_data.get("grocery_link", "N/A")
            }
            for category in grocery_data.get("grocery_details", {}).get("categories", []):
                for sub_category in category.get("sub_categories", []):
                    for item in sub_category.get("Items", []):
                        simplified_data.append({
                            **general_info,
                            "Category": category.get("name", "N/A"),
                            "Sub-Category": sub_category.get("sub_category_name", "N/A"),
                            "Item Name": item.get("item_name", "N/A"),
                            "Item Price": item.get("item_price", "N/A"),
                            "Item Description": item.get("item_description", "N/A"),
                            "Item Link": item.get("item_link", "N/A")
                        })
        logging.debug(f"Simplified data for Excel: {simplified_data}")
        if simplified_data:
            df = pd.DataFrame(simplified_data)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        else:
            logging.warning(f"No data to write to Excel for sheet: {sheet_name}")

    @retry(tries=3, delay=2, backoff=2)
    def upload_to_drive(self, file_path):
        logging.info(f"Uploading {file_path} to Google Drive...")
        try:
            # Verify credentials are available
            if not self.drive_uploader.credentials_json:
                logging.error("TALABAT_GCLOUD_KEY_JSON is not set or invalid.")
                return False
            # Authenticate
            if not self.drive_uploader.authenticate():
                logging.error("Failed to authenticate with Google Drive. Check TALABAT_GCLOUD_KEY_JSON validity.")
                return False
            # Upload file
            file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
            success = len(file_ids) == 2
            if success:
                logging.info(f"Successfully uploaded {file_path} to Google Drive")
            else:
                logging.error(f"Failed to upload {file_path}: Incomplete upload to folders")
            return success
        except Exception as e:
            if "429" in str(e) or "rate limit" in str(e).lower():
                logging.error(f"Rate limit exceeded while uploading to Google Drive: {str(e)}")
            else:
                logging.error(f"Error uploading to Google Drive: {str(e)}")
            return False

    async def run(self):
        ahmadi_areas = [
            ("الظهر", "https://www.talabat.com/kuwait/groceries/59/dhaher"),
        ]
        excel_filename = os.path.join(self.output_dir, "الاحمدي_groceries.xlsx")
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            for idx, (area_name, area_url) in enumerate(ahmadi_areas):
                self.current_progress["current_area_index"] = idx
                self.scraped_progress["current_area_index"] = idx
                area_results = await self.scrape_and_save_area(area_name, area_url, browser)
                self.create_excel_sheet(workbook, area_name, self.scraped_progress["all_results"][area_name])
                workbook.save(excel_filename)
                self.current_progress["completed_areas"].append(area_name)
                self.scraped_progress["completed_areas"].append(area_name)
                self.save_current_progress()
                self.save_scraped_progress()
                self.commit_progress(f"Completed {area_name}")
            await browser.close()

        print(f"SCRAPING COMPLETED\nExcel file saved: {excel_filename}")
        self.upload_to_drive(excel_filename)
        print("Uploaded Excel to Google Drive")

async def main():
    scraper = MainScraper()
    await scraper.run()

if __name__ == "__main__":
    asyncio.run(main())







# import asyncio
# import json
# import os
# import tempfile
# import sys
# import subprocess
# import re
# from typing import Dict, List
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
# from SavingOnDrive import SavingOnDrive
# import logging
# from datetime import datetime

# # Set up logging (used by MainScraper)
# logging.basicConfig(
#     filename='scraper.log',
#     level=logging.DEBUG,
#     format='%(asctime)s - %(levelname)s - %(message)s'
# )

# class TalabatGroceries:
#     def __init__(self, url, browser):
#         self.url = url
#         self.base_url = "https://www.talabat.com"
#         self.browser = browser
#         print(f"Initialized TalabatGroceries with URL: {self.url}")

#     async def get_general_link(self, page):
#         print("Attempting to get general link")
#         retries = 3
#         while retries > 0:
#             try:
#                 link_element = await page.wait_for_selector('//a[@data-testid="view-all-link"]', timeout=60000)
#                 if link_element:
#                     full_link = self.base_url + await link_element.get_attribute('href')
#                     print(f"General link found: {full_link}")
#                     return full_link
#                 else:
#                     print("General link not found")
#                     return None
#             except Exception as e:
#                 print(f"Error getting general link: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return None

#     async def get_delivery_fees(self, page):
#         print("Attempting to get delivery fees")
#         retries = 3
#         while retries > 0:
#             try:
#                 delivery_fees_element = await page.query_selector('xpath=/html/body/div/div/div[1]/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[2]/span[1]')
#                 delivery_fees = await delivery_fees_element.inner_text() if delivery_fees_element else "N/A"
#                 print(f"Delivery fees: {delivery_fees}")
#                 return delivery_fees
#             except Exception as e:
#                 print(f"Error getting delivery fees: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return "N/A"

#     async def get_minimum_order(self, page):
#         print("Attempting to get minimum order")
#         retries = 3
#         while retries > 0:
#             try:
#                 minimum_order_element = await page.query_selector('xpath=/html/body/div/div/div[1]/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[2]/span[3]')
#                 minimum_order = await minimum_order_element.inner_text() if minimum_order_element else "N/A"
#                 print(f"Minimum order: {minimum_order}")
#                 return minimum_order
#             except Exception as e:
#                 print(f"Error getting minimum order: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return "N/A"

#     async def extract_category_names(self, page):
#         print("Attempting to extract category names")
#         retries = 3
#         while retries > 0:
#             try:
#                 category_name_elements = await page.query_selector_all('//span[@data-testid="category-name"]')
#                 category_names = [await element.inner_text() for element in category_name_elements]
#                 print(f"Category names extracted: {category_names}")
#                 return category_names
#             except Exception as e:
#                 print(f"Error extracting category names: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return []

#     async def extract_category_links(self, page):
#         print("Attempting to extract category links")
#         retries = 3
#         while retries > 0:
#             try:
#                 category_link_elements = await page.query_selector_all('//a[@data-testid="category-item-container"]')
#                 category_links = [self.base_url + await element.get_attribute('href') for element in category_link_elements]
#                 print(f"Category links extracted: {category_links}")
#                 return category_links
#             except Exception as e:
#                 print(f"Error extracting category links: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return []

#     async def extract_sub_categories(self, page, category_link):
#         print(f"Attempting to extract sub-categories for: {category_link}")
#         retries = 3
#         while retries > 0:
#             try:
#                 await page.goto(category_link, timeout=240000)
#                 await page.wait_for_load_state("networkidle", timeout=240000)
#                 sub_category_elements = await page.query_selector_all('//div[@data-test="sub-category-container"]//a[@data-testid="subCategory-a"]')
#                 sub_categories = []
#                 for element in sub_category_elements:  # Removed [:5] limit
#                     try:
#                         sub_category_name = await element.inner_text()
#                         sub_category_link = self.base_url + await element.get_attribute('href')
#                         print(f"    Processing sub-category: {sub_category_name}")
#                         print(f"    Sub-category link: {sub_category_link}")
#                         items = await self.extract_all_items_from_sub_category(sub_category_link)
#                         sub_categories.append({
#                             "sub_category_name": sub_category_name,
#                             "sub_category_link": sub_category_link,
#                             "Items": items
#                         })
#                     except Exception as e:
#                         print(f"Error processing sub-category: {e}")
#                 print(f"Extracted {len(sub_categories)} sub-categories")
#                 return sub_categories
#             except Exception as e:
#                 print(f"Error extracting sub-categories: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return []

#     async def extract_item_details(self, item_link):
#         print(f"Attempting to extract item details for link: {item_link}")
#         retries = 3
#         while retries > 0:
#             try:
#                 page = await self.browser.new_page()
#                 await page.goto(item_link, timeout=240000)
#                 await page.wait_for_load_state("networkidle", timeout=240000)
#                 # Add a small delay and scroll to ensure dynamic content loads
#                 await page.wait_for_timeout(5000)  # Wait 5 seconds
#                 await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")  # Scroll to bottom
#                 await page.wait_for_load_state("networkidle", timeout=240000)
    
#                 # Try multiple price selectors
#                 price_selectors = [
#                     '//div[@class="price"]//span[@class="currency "]',
#                     '//span[contains(@class, "price")]',
#                     '//div[contains(@class, "price")]//text()'
#                 ]
#                 item_price = "N/A"
#                 for selector in price_selectors:
#                     price_element = await page.query_selector(selector)
#                     if price_element:
#                         item_price = await price_element.inner_text()
#                         break
    
#                 # Try multiple description selectors
#                 desc_selectors = [
#                     '//div[@class="description"]//p[@data-testid="item-description"]',
#                     '//div[contains(@class, "description")]//p',
#                     '//p[contains(@class, "description")]'
#                 ]
#                 item_description = "N/A"
#                 for selector in desc_selectors:
#                     desc_element = await page.query_selector(selector)
#                     if desc_element:
#                         item_description = await desc_element.inner_text()
#                         break
    
#                 # Delivery time selector
#                 delivery_time_element = await page.query_selector('//div[@data-testid="delivery-tag"]//span')
#                 delivery_time = await delivery_time_element.inner_text() if delivery_time_element else "N/A"
    
#                 # Image selector with fallback
#                 image_selectors = [
#                     '//div[@data-testid="item-image"]//img',
#                     '//img[contains(@class, "item-image")]',
#                     '//img[@alt="product image"]'
#                 ]
#                 item_images = []
#                 for selector in image_selectors:
#                     item_image_elements = await page.query_selector_all(selector)
#                     if item_image_elements:
#                         item_images = [await img.get_attribute('src') for img in item_image_elements if await img.get_attribute('src')]
#                         break
    
#                 # Print in the desired format
#                 print(f"Item price: {item_price}")
#                 print(f"Item description: {item_description}")
#                 print(f"Delivery time range: {delivery_time}")
#                 print(f"Item images: {item_images}")
    
#                 await page.close()
#                 return {
#                     "item_price": item_price,
#                     "item_description": item_description,
#                     "item_delivery_time_range": delivery_time,
#                     "item_images": item_images
#                 }
#             except Exception as e:
#                 print(f"Error extracting item details for {item_link}: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         print(f"Failed to extract details for {item_link} after all retries")
#         return {
#             "item_price": "N/A",
#             "item_description": "N/A",
#             "item_delivery_time_range": "N/A",
#             "item_images": []
#         } 
#     async def extract_all_items_from_sub_category(self, sub_category_link):
#         print(f"Attempting to extract all items from sub-category: {sub_category_link}")
#         retries = 3
#         while retries > 0:
#             try:
#                 sub_page = await self.browser.new_page()
#                 await sub_page.goto(sub_category_link, timeout=240000)
#                 await sub_page.wait_for_load_state("networkidle", timeout=240000)
#                 await sub_page.wait_for_selector('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]', timeout=240000)

#                 pagination_element = await sub_page.query_selector('//div[@class="sc-104fa483-0 fCcIDQ"]//ul[@class="paginate-wrap"]')
#                 total_pages = 1
#                 if pagination_element:
#                     page_numbers = await pagination_element.query_selector_all('//li[contains(@class, "paginate-li f-16 f-500")]//a')
#                     total_pages = len(page_numbers) if page_numbers else 1
#                 print(f"      Found {total_pages} pages in this sub-category")

#                 items = []
#                 for page_number in range(1, total_pages + 1):  # Removed page limit
#                     print(f"      Processing page {page_number} of {total_pages}")
#                     page_url = f"{sub_category_link}&page={page_number}"
#                     await sub_page.goto(page_url, timeout=240000)
#                     await sub_page.wait_for_load_state("networkidle", timeout=240000)
#                     await sub_page.wait_for_selector('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]', timeout=240000)

#                     item_elements = await sub_page.query_selector_all('//div[@class="category-items-container all-items w-100"]//div[@class="col-8 col-sm-4"]//a[@data-testid="grocery-item-link-nofollow"]')
#                     print(f"        Found {len(item_elements)} items on page {page_number}")

#                     for i, element in enumerate(item_elements):  # Removed [:10] limit
#                         try:
#                             item_name_element = await element.query_selector('div[data-test="item-name"]')
#                             item_name = await item_name_element.inner_text() if item_name_element else f"Unknown Item {i+1}"
#                             print(f"        Item name: {item_name}")

#                             item_link = self.base_url + await element.get_attribute('href')
#                             print(f"        Item link: {item_link}")

#                             item_details = await self.extract_item_details(item_link)
#                             items.append({
#                                 "item_name": item_name,
#                                 "item_link": item_link,
#                                 **item_details
#                             })
#                         except Exception as e:
#                             print(f"        Error processing item {i+1}: {e}")
#                 await sub_page.close()
#                 return items
#             except Exception as e:
#                 print(f"Error extracting items from sub-category {sub_category_link}: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return []

#     async def extract_categories(self, page):
#         print(f"Processing grocery: {self.url}")
#         retries = 3
#         while retries > 0:
#             try:
#                 await page.goto(self.url, timeout=240000)
#                 await page.wait_for_load_state("networkidle", timeout=240000)
#                 print("Page loaded successfully")

#                 delivery_fees = await self.get_delivery_fees(page)
#                 minimum_order = await self.get_minimum_order(page)
#                 view_all_link = await self.get_general_link(page)

#                 print(f"  Delivery fees: {delivery_fees}")
#                 print(f"  Minimum order: {minimum_order}")

#                 categories_data = []
#                 if view_all_link:
#                     print(f"  Navigating to view all link: {view_all_link}")
#                     category_page = await self.browser.new_page()
#                     await category_page.goto(view_all_link, timeout=240000)
#                     await category_page.wait_for_load_state("networkidle", timeout=240000)

#                     category_names = await self.extract_category_names(category_page)
#                     category_links = await self.extract_category_links(category_page)

#                     print(f"  Found {len(category_names)} categories")

#                     for index, (name, link) in enumerate(zip(category_names, category_links)):  # Removed [:3] limit
#                         print(f"  Processing category {index+1}/{len(category_names)}: {name}")
#                         print(f"  Category link: {link}")
#                         sub_category_page = await self.browser.new_page()
#                         sub_categories = await self.extract_sub_categories(sub_category_page, link)
#                         await sub_category_page.close()
#                         print(f"  Found {len(sub_categories)} sub-categories in {name}")
#                         categories_data.append({
#                             "name": name,
#                             "link": link,
#                             "sub_categories": sub_categories
#                         })
#                     await category_page.close()

#                 return {
#                     "delivery_fees": delivery_fees,
#                     "minimum_order": minimum_order,
#                     "categories": categories_data
#                 }
#             except Exception as e:
#                 print(f"Error extracting categories: {e}")
#                 retries -= 1
#                 print(f"Retries left: {retries}")
#                 await asyncio.sleep(5)
#         return {"error": "Failed to extract categories after multiple attempts"}

# class MainScraper:
#     CURRENT_PROGRESS_FILE = "current_progress.json"
#     SCRAPED_PROGRESS_FILE = "scraped_progress.json"

#     def __init__(self):
#         self.output_dir = "output"
#         self.drive_uploader = SavingOnDrive('credentials.json')
#         os.makedirs(self.output_dir, exist_ok=True)
#         self.current_progress = self.load_current_progress()
#         self.scraped_progress = self.load_scraped_progress()
#         self.github_token = os.environ.get('GITHUB_TOKEN')
#         self.ensure_playwright_browsers()
#         # Force reset completed_areas to ensure scraping
#         self.current_progress["completed_areas"] = []
#         self.scraped_progress["completed_areas"] = []
#         self.save_current_progress()
#         self.save_scraped_progress()

#     def ensure_playwright_browsers(self):
#         subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)

#     def load_current_progress(self) -> Dict:
#         if os.path.exists(self.CURRENT_PROGRESS_FILE):
#             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 return json.load(f)
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "current_progress": {
#                 "area_name": None,
#                 "current_grocery": 0,
#                 "total_groceries": 0,
#                 "processed_groceries": [],
#                 "current_category": None,
#                 "current_sub_category": None,
#                 "completed_categories": {}
#             }
#         }
#         self.save_current_progress(default_progress)
#         return default_progress

#     def save_current_progress(self, progress: Dict = None):
#         progress = progress or self.current_progress
#         progress["last_updated"] = datetime.now().isoformat()
#         with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#             json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#             os.replace(temp_file.name, self.CURRENT_PROGRESS_FILE)
#         print(f"Saved progress to {self.CURRENT_PROGRESS_FILE}")

#     def load_scraped_progress(self) -> Dict:
#         if os.path.exists(self.SCRAPED_PROGRESS_FILE):
#             with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 return json.load(f)
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "all_results": {},
#             "current_progress": self.current_progress["current_progress"]
#         }
#         self.save_scraped_progress(default_progress)
#         return default_progress

#     def save_scraped_progress(self, progress: Dict = None):
#         progress = progress or self.scraped_progress
#         progress["last_updated"] = datetime.now().isoformat()
#         with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#             json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#             os.replace(temp_file.name, self.SCRAPED_PROGRESS_FILE)
#         print(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE}")

#     def commit_progress(self, message: str = "Periodic progress update"):
#         if not self.github_token:
#             print("No GITHUB_TOKEN, skipping commit")
#             return
#         subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE, self.SCRAPED_PROGRESS_FILE, self.output_dir], check=True)
#         result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
#         if result.returncode == 0:
#             subprocess.run(["git", "push"], check=True, env={"GIT_AUTH_TOKEN": self.github_token})
#             print(f"Committed progress: {message}")

#     async def scrape_and_save_area(self, area_name: str, area_url: str, browser) -> List[Dict]:
#         print(f"\n{'='*50}\nSCRAPING AREA: {area_name}\nURL: {area_url}\n{'='*50}")
#         all_area_results = self.scraped_progress["all_results"].get(area_name, {})
#         current_progress = self.current_progress["current_progress"]
#         scraped_current_progress = self.scraped_progress["current_progress"]

#         is_resuming = current_progress["area_name"] == area_name
#         start_grocery = current_progress["current_grocery"] if is_resuming else 0

#         if not is_resuming:
#             current_progress.update({
#                 "area_name": area_name,
#                 "current_grocery": 0,
#                 "total_groceries": 0,
#                 "processed_groceries": [],
#                 "current_category": None,
#                 "current_sub_category": None,
#                 "completed_categories": {}
#             })
#             scraped_current_progress.update(current_progress)
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Started scraping {area_name}")

#         page = await browser.new_page()
#         await page.goto(area_url, timeout=60000)
#         groceries_on_page = await self.get_page_groceries(page)
#         current_progress["total_groceries"] = len(groceries_on_page)
#         scraped_current_progress["total_groceries"] = len(groceries_on_page)
#         print(f"Found {len(groceries_on_page)} groceries")
#         await page.close()

#         for grocery_idx, grocery in enumerate(groceries_on_page):
#             grocery_num = grocery_idx + 1
#             if grocery_num <= start_grocery:
#                 continue
#             grocery_title = grocery["grocery_title"]
#             if grocery_title in current_progress["processed_groceries"]:
#                 continue

#             current_progress["current_grocery"] = grocery_num
#             scraped_current_progress["current_grocery"] = grocery_num
#             print(f"Processing grocery {grocery_num}/{len(groceries_on_page)}: {grocery_title}")

#             grocery_page = await browser.new_page()
#             talabat_grocery = TalabatGroceries(grocery["grocery_link"], browser)
#             grocery_details = await talabat_grocery.extract_categories(grocery_page)
#             all_area_results[grocery_title] = {
#                 "grocery_link": grocery["grocery_link"],
#                 "delivery_time": grocery["delivery_time"],
#                 "grocery_details": grocery_details
#             }
#             await self.process_grocery_categories(grocery_title, grocery_details, talabat_grocery, grocery_page)
#             await grocery_page.close()

#             current_progress["processed_groceries"].append(grocery_title)
#             scraped_current_progress["processed_groceries"].append(grocery_title)
#             self.scraped_progress["all_results"][area_name] = all_area_results
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Processed {grocery_title}")

#         json_filename = os.path.join(self.output_dir, f"{area_name}.json")
#         with open(json_filename, 'w', encoding='utf-8') as f:
#             json.dump(all_area_results, f, indent=2)

#         workbook = Workbook()
#         self.create_excel_sheet(workbook, area_name, all_area_results)
#         excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
#         workbook.save(excel_filename)
#         self.upload_to_drive(excel_filename)

#         current_progress.update({"area_name": None, "current_grocery": 0, "total_groceries": 0, "processed_groceries": []})
#         scraped_current_progress.update(current_progress)
#         self.save_current_progress()
#         self.save_scraped_progress()
#         self.commit_progress(f"Completed {area_name}")
#         return list(all_area_results.values())

#     async def process_grocery_categories(self, grocery_title, grocery_details, talabat_grocery, page):
#         completed_categories = self.current_progress["current_progress"]["completed_categories"].get(grocery_title, {})
#         for category in grocery_details.get("categories", []):
#             category_name = category["name"]
#             if category_name not in completed_categories:
#                 await self.process_category(grocery_title, category, talabat_grocery, page)

#     async def process_category(self, grocery_title, category_info, talabat_grocery, page):
#         category_name = category_info["name"]
#         self.current_progress["current_progress"]["current_category"] = category_name
#         self.scraped_progress["current_progress"]["current_category"] = category_name
#         sub_categories = await talabat_grocery.extract_sub_categories(page, category_info["link"])
#         category_info["sub_categories"] = sub_categories
#         self.current_progress["current_progress"]["completed_categories"].setdefault(grocery_title, {})[category_name] = [sub["sub_category_name"] for sub in sub_categories]
#         self.scraped_progress["current_progress"]["completed_categories"].setdefault(grocery_title, {})[category_name] = [sub["sub_category_name"] for sub in sub_categories]
#         self.save_current_progress()
#         self.save_scraped_progress()

#     async def get_page_groceries(self, page) -> List[Dict]:
#         logging.info("Extracting grocery information")
#         try:
#             await page.wait_for_selector('div[data-testid="one-vendor-container"]', timeout=30000)
#             vendor_containers = await page.query_selector_all('div[data-testid="one-vendor-container"]')
#             groceries_info = []
#             for container in vendor_containers:
#                 title_element = await container.query_selector('a div h2')
#                 title = await title_element.inner_text() if title_element else "Unknown Grocery"
#                 link_element = await container.query_selector('a')
#                 link = "https://www.talabat.com" + await link_element.get_attribute('href') if link_element else None
#                 delivery_info = await container.query_selector('div.deliveryInfo')
#                 delivery_time_text = await delivery_info.inner_text() if delivery_info else ""
#                 delivery_time = re.findall(r'\d+', delivery_time_text)[0] + " mins" if re.findall(r'\d+', delivery_time_text) else "N/A"
#                 if link:
#                     groceries_info.append({"grocery_title": title, "grocery_link": link, "delivery_time": delivery_time})
#             logging.info(f"Extracted {len(groceries_info)} groceries: {[g['grocery_title'] for g in groceries_info]}")
#             return groceries_info
#         except Exception as e:
#             logging.error(f"Error extracting groceries: {e}")
#             return []

#     def create_excel_sheet(self, workbook, sheet_name: str, data: Dict):
#         sheet = workbook.create_sheet(title=sheet_name)
#         simplified_data = []
#         for grocery_title, grocery_data in data.items():
#             general_info = {
#                 "Grocery Title": grocery_title,
#                 "Delivery Time": grocery_data.get("delivery_time", "N/A"),
#                 "Delivery Fees": grocery_data.get("grocery_details", {}).get("delivery_fees", "N/A"),
#                 "Minimum Order": grocery_data.get("grocery_details", {}).get("minimum_order", "N/A"),
#                 "URL": grocery_data.get("grocery_link", "N/A")
#             }
#             for category in grocery_data.get("grocery_details", {}).get("categories", []):
#                 for sub_category in category.get("sub_categories", []):
#                     for item in sub_category.get("Items", []):
#                         simplified_data.append({
#                             **general_info,
#                             "Category": category.get("name", "N/A"),
#                             "Sub-Category": sub_category.get("sub_category_name", "N/A"),
#                             "Item Name": item.get("item_name", "N/A"),
#                             "Item Price": item.get("item_price", "N/A"),
#                             "Item Description": item.get("item_description", "N/A"),
#                             "Item Link": item.get("item_link", "N/A")
#                         })
#         logging.debug(f"Simplified data for Excel: {simplified_data}")
#         if simplified_data:
#             df = pd.DataFrame(simplified_data)
#             for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     sheet.cell(row=r_idx, column=c_idx, value=value)
#         else:
#             logging.warning(f"No data to write to Excel for sheet: {sheet_name}")

#     def upload_to_drive(self, file_path):
#         print(f"Uploading {file_path} to Google Drive...")
#         try:
#             file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
#             print(f"Uploaded {file_path} successfully" if len(file_ids) == 2 else "Upload failed")
#         except Exception as e:
#             print(f"Error uploading to Drive: {e}")

#     async def run(self):
#         ahmadi_areas = [
#             ("الظهر", "https://www.talabat.com/kuwait/groceries/59/dhaher"),
#         ]
#         excel_filename = os.path.join(self.output_dir, "الاحمدي_groceries.xlsx")
#         workbook = Workbook()
#         if "Sheet" in workbook.sheetnames:
#             workbook.remove(workbook["Sheet"])

#         async with async_playwright() as p:
#             browser = await p.chromium.launch(headless=True)
#             for idx, (area_name, area_url) in enumerate(ahmadi_areas):
#                 self.current_progress["current_area_index"] = idx
#                 self.scraped_progress["current_area_index"] = idx
#                 area_results = await self.scrape_and_save_area(area_name, area_url, browser)
#                 self.create_excel_sheet(workbook, area_name, self.scraped_progress["all_results"][area_name])
#                 workbook.save(excel_filename)
#                 self.current_progress["completed_areas"].append(area_name)
#                 self.scraped_progress["completed_areas"].append(area_name)
#                 self.save_current_progress()
#                 self.save_scraped_progress()
#                 self.commit_progress(f"Completed {area_name}")
#             await browser.close()

#         print(f"SCRAPING COMPLETED\nExcel file saved: {excel_filename}")
#         self.upload_to_drive(excel_filename)
#         print("Uploaded Excel to Google Drive")

# def create_credentials_file():
#     credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
#     if credentials_json:
#         with open('credentials.json', 'w') as f:
#             f.write(credentials_json)
#         return True
#     print("ERROR: TALABAT_GCLOUD_KEY_JSON not set!")
#     return False

# async def main():
#     if not create_credentials_file():
#         sys.exit(1)
#     scraper = MainScraper()
#     await scraper.run()

# if __name__ == "__main__":
#     asyncio.run(main())



# # import asyncio
# # import json
# # import os
# # import re
# # import nest_asyncio
# # from bs4 import BeautifulSoup
# # import pandas as pd
# # from playwright.async_api import async_playwright
# # from SavingOnDrive import SavingOnDrive

# # class MainScraper:
# #     def __init__(self, target_url="https://www.talabat.com/kuwait/groceries/59/dhaher"):
# #         self.target_url = target_url
# #         self.base_url = "https://www.talabat.com"
# #         self.json_file = "talabat_groceries.json"
# #         self.excel_file = "dhaher.xlsx"
# #         self.groceries_data = {}
# #         self.drive_uploader = SavingOnDrive('credentials.json')  # Initialize SavingOnDrive
# #         print(f"Initialized MainScraper with target URL: {self.target_url}")

# #         if not os.path.exists(self.json_file):
# #             print(f"Creating new JSON file: {self.json_file}")
# #             with open(self.json_file, 'w') as f:
# #                 json.dump({}, f)
# #         else:
# #             print(f"Loading data from JSON file: {self.json_file}")
# #             with open(self.json_file, 'r') as f:
# #                 try:
# #                     self.groceries_data = json.load(f)
# #                     print(f"Loaded data: {self.groceries_data}")
# #                 except json.JSONDecodeError:
# #                     print("Error decoding JSON file. Starting with an empty dictionary.")
# #                     self.groceries_data = {}

# #     async def extract_grocery_info(self, page):
# #         print("Attempting to extract grocery information")
# #         retries = 3
# #         while retries > 0:
# #             try:
# #                 vendor_containers = await page.query_selector_all('div[data-testid="one-vendor-container"]')
# #                 print(f"Found {len(vendor_containers)} grocery vendors on page")
# #                 groceries_info = []
# #                 for i, container in enumerate(vendor_containers, 1):
# #                     try:
# #                         title_element = await container.query_selector('a div h2')
# #                         grocery_title = await title_element.inner_text() if title_element else f"Unknown Grocery {i}"
# #                         print(f"  Grocery title: {grocery_title}")
# #                         link_element = await container.query_selector('a')
# #                         grocery_link = self.base_url + await link_element.get_attribute('href') if link_element else None
# #                         print(f"  Grocery link: {grocery_link}")
# #                         delivery_info = await container.query_selector('div.deliveryInfo')
# #                         delivery_time_text = await delivery_info.inner_text() if delivery_info else "N/A"
# #                         print(f"  Delivery time text: {delivery_time_text}")
# #                         if delivery_time_text != "N/A":
# #                             digits = re.findall(r'\d+', delivery_time_text)
# #                             delivery_time = f"{digits[0]} mins" if digits else "N/A"
# #                         else:
# #                             delivery_time = "N/A"
# #                         print(f"  Delivery time: {delivery_time}")
# #                         if grocery_title and grocery_link and delivery_time:
# #                             print(f"Found grocery: {grocery_title}, Delivery time: {delivery_time}")
# #                             groceries_info.append({
# #                                 'grocery_title': grocery_title,
# #                                 'grocery_link': grocery_link,
# #                                 'delivery_time': delivery_time
# #                             })
# #                     except Exception as e:
# #                         print(f"Error processing grocery {i}: {e}")
# #                 return groceries_info
# #             except Exception as e:
# #                 print(f"Error extracting grocery information: {e}")
# #                 retries -= 1
# #                 print(f"Retries left: {retries}")
# #                 await asyncio.sleep(5)
# #         return []

# #     def save_to_json(self):
# #         print("Saving data to JSON file")
# #         try:
# #             with open(self.json_file, 'w') as f:
# #                 json.dump(self.groceries_data, f, indent=4)
# #             print(f"Data saved to {self.json_file}")
# #         except Exception as e:
# #             print(f"Error saving to JSON file: {e}")
               
# #     def save_to_excel(self):
# #         print("Saving data to Excel file")
# #         try:
# #             writer = pd.ExcelWriter(self.excel_file, engine='xlsxwriter')
# #             for grocery_title, grocery_data in self.groceries_data.items():
# #                 flattened_data = []
# #                 general_info = {
# #                     'grocery_title': grocery_title,
# #                     'delivery_time': grocery_data.get('delivery_time', 'N/A'),
# #                     'delivery_fees': grocery_data.get('grocery_details', {}).get('delivery_fees', 'N/A'),
# #                     'minimum_order': grocery_data.get('grocery_details', {}).get('minimum_order', 'N/A')
# #                 }
# #                 categories = grocery_data.get('grocery_details', {}).get('categories', [])
# #                 for category in categories:
# #                     category_name = category.get('name', 'N/A')
# #                     for sub_category in category.get('sub_categories', []):
# #                         sub_category_name = sub_category.get('sub_category_name', 'N/A')
# #                         for item in sub_category.get('Items', []):
# #                             item_data = {
# #                                 **general_info,
# #                                 'category': category_name,
# #                                 'sub_category': sub_category_name,
# #                                 'item_name': item.get('item_name', 'N/A'),
# #                                 'item_price': item.get('item_price', 'N/A'),
# #                                 'item_description': item.get('item_description', 'N/A'),
# #                                 'item_link': item.get('item_link', 'N/A')
# #                             }
# #                             flattened_data.append(item_data)
# #                 if not flattened_data:
# #                     flattened_data.append(general_info)
# #                 safe_title = re.sub(r'[\\/*?:\[\]]', '_', grocery_title)[:31]
# #                 df = pd.DataFrame(flattened_data)
# #                 df.to_excel(writer, sheet_name=safe_title, index=False)
# #             writer.close()
# #             print(f"Data saved to {self.excel_file}")
# #             # Upload to Google Drive after saving
# #             self.upload_to_drive()
# #         except Exception as e:
# #             print(f"Error saving to Excel: {e}")

# #     def upload_to_drive(self):
# #         """Upload the Excel file to Google Drive"""
# #         print(f"\nUploading {self.excel_file} to Google Drive...")
# #         try:
# #             if not self.drive_uploader.authenticate():
# #                 print("Failed to authenticate with Google Drive")
# #                 return False
# #             file_ids = self.drive_uploader.upload_to_multiple_folders(self.excel_file)
# #             if len(file_ids) == len(self.drive_uploader.target_folders):
# #                 print(f"Uploaded {self.excel_file} to Google Drive successfully")
# #                 return True
# #             else:
# #                 print(f"Failed to upload {self.excel_file} to all target folders")
# #                 return False
# #         except Exception as e:
# #             print(f"Error uploading to Google Drive: {str(e)}")
# #             return False

# #     async def process_grocery(self, grocery_info):
# #         grocery_title = grocery_info['grocery_title']
# #         grocery_link = grocery_info['grocery_link']
# #         delivery_time = grocery_info['delivery_time']
# #         print(f"Processing grocery: {grocery_title}")
# #         if grocery_title in self.groceries_data:
# #             print(f"Skipping {grocery_title} - already processed")
# #             return
# #         talabat_grocery = TalabatGroceries(grocery_link)
# #         for browser_type in ["chromium", "firefox"]:
# #             try:
# #                 async with async_playwright() as p:
# #                     browser = await p[browser_type].launch(headless=True)
# #                     page = await browser.new_page()
# #                     grocery_details = await talabat_grocery.extract_categories(page)
# #                     self.groceries_data[grocery_title] = {
# #                         'grocery_link': grocery_link,
# #                         'delivery_time': delivery_time,
# #                         'grocery_details': grocery_details
# #                     }
# #                     self.save_to_json()
# #                     print(f"Successfully processed and saved data for {grocery_title}")
# #                     break
# #             except Exception as e:
# #                 print(f"Error processing {grocery_title} with {browser_type}: {e}")
# #                 continue
# #             finally:
# #                 await browser.close()

# #     async def run(self):
# #         print("Starting the scraper")
# #         retries = 3
# #         while retries > 0:
# #             try:
# #                 print(f"Target URL: {self.target_url}")
# #                 async with async_playwright() as p:
# #                     browser = await p.chromium.launch(headless=True)
# #                     page = await browser.new_page()
# #                     page.set_default_timeout(240000)
# #                     await page.goto(self.target_url, timeout=240000)
# #                     await page.wait_for_load_state("networkidle", timeout=240000)
# #                     print("Page loaded successfully")
# #                     try:
# #                         await page.wait_for_selector('div[data-testid="one-vendor-container"]', timeout=240000)
# #                         print("Grocery vendor elements found")
# #                     except Exception as e:
# #                         print(f"Error waiting for vendor elements: {e}")
# #                         CHIprint("Attempting to continue anyway...")
# #                     groceries_info = await self.extract_grocery_info(page)
# #                     await browser.close()
# #                     print(f"Found {len(groceries_info)} groceries to process")
# #                     for grocery_info in groceries_info:
# #                         await self.process_grocery(grocery_info)
# #                     self.save_to_excel()
# #                     print("Scraping completed successfully")
# #                 break
# #             except Exception as e:
# #                 print(f"Error in main scraper: {e}")
# #                 retries -= 1
# #                 print(f"Retries left: {retries}")
# #                 await asyncio.sleep(5)
# #                 if retries == 0:
# #                     print("Failed to complete the scraping process after multiple attempts.")

# # async def main():
# #     if not os.path.exists('credentials.json'):
# #         credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
# #         if credentials_json:
# #             with open('credentials.json', 'w') as f:
# #                 f.write(credentials_json)
# #             print("Created credentials.json from environment variable")
# #         else:
# #             print("ERROR: credentials.json not found and TALABAT_GCLOUD_KEY_JSON not set!")
# #             return
# #     scraper = MainScraper()
# #     await scraper.run()

# # if __name__ == "__main__":
# #     asyncio.run(main())
# # else:
# #     asyncio.get_event_loop().run_until_complete(main())

# # # class MainScraper:
# # #     def __init__(self, target_url="https://www.talabat.com/kuwait/groceries/59/dhaher"):
# # #         self.target_url = target_url
# # #         self.base_url = "https://www.talabat.com"
# # #         self.json_file = "talabat_groceries.json"
# # #         self.excel_file = "dhaher.xlsx"
# # #         self.groceries_data = {}
# # #         print(f"Initialized MainScraper with target URL: {self.target_url}")

# # #         if not os.path.exists(self.json_file):
# # #             print(f"Creating new JSON file: {self.json_file}")
# # #             with open(self.json_file, 'w') as f:
# # #                 json.dump({}, f)
# # #         else:
# # #             print(f"Loading data from JSON file: {self.json_file}")
# # #             with open(self.json_file, 'r') as f:
# # #                 try:
# # #                     self.groceries_data = json.load(f)
# # #                     print(f"Loaded data: {self.groceries_data}")
# # #                 except json.JSONDecodeError:
# # #                     print("Error decoding JSON file. Starting with an empty dictionary.")
# # #                     self.groceries_data = {}

# # #     async def extract_grocery_info(self, page):
# # #         print("Attempting to extract grocery information")
# # #         retries = 3
# # #         while retries > 0:
# # #             try:
# # #                 vendor_containers = await page.query_selector_all('div[data-testid="one-vendor-container"]')
# # #                 print(f"Found {len(vendor_containers)} grocery vendors on page")

# # #                 groceries_info = []
# # #                 for i, container in enumerate(vendor_containers, 1):
# # #                     try:
# # #                         title_element = await container.query_selector('a div h2')
# # #                         grocery_title = await title_element.inner_text() if title_element else f"Unknown Grocery {i}"
# # #                         print(f"  Grocery title: {grocery_title}")

# # #                         link_element = await container.query_selector('a')
# # #                         grocery_link = self.base_url + await link_element.get_attribute('href') if link_element else None
# # #                         print(f"  Grocery link: {grocery_link}")

# # #                         delivery_info = await container.query_selector('div.deliveryInfo')
# # #                         delivery_time_text = await delivery_info.inner_text() if delivery_info else "N/A"
# # #                         print(f"  Delivery time text: {delivery_time_text}")

# # #                         if delivery_time_text != "N/A":
# # #                             digits = re.findall(r'\d+', delivery_time_text)
# # #                             delivery_time = f"{digits[0]} mins" if digits else "N/A"
# # #                         else:
# # #                             delivery_time = "N/A"
# # #                         print(f"  Delivery time: {delivery_time}")

# # #                         if grocery_title and grocery_link and delivery_time:
# # #                             print(f"Found grocery: {grocery_title}, Delivery time: {delivery_time}")
# # #                             groceries_info.append({
# # #                                 'grocery_title': grocery_title,
# # #                                 'grocery_link': grocery_link,
# # #                                 'delivery_time': delivery_time
# # #                             })
# # #                     except Exception as e:
# # #                         print(f"Error processing grocery {i}: {e}")
# # #                 return groceries_info
# # #             except Exception as e:
# # #                 print(f"Error extracting grocery information: {e}")
# # #                 retries -= 1
# # #                 print(f"Retries left: {retries}")
# # #                 await asyncio.sleep(5)
# # #         return []

# # #     def save_to_json(self):
# # #         """Save scraped data to JSON file"""
# # #         print("Saving data to JSON file")
# # #         try:
# # #             with open(self.json_file, 'w') as f:
# # #                 json.dump(self.groceries_data, f, indent=4)
# # #             print(f"Data saved to {self.json_file}")
# # #         except Exception as e:
# # #             print(f"Error saving to JSON file: {e}")
               
# # #     def save_to_excel(self):
# # #         """Save data from JSON to Excel file with each grocery in a separate sheet"""
# # #         print("Saving data to Excel file")
# # #         try:
# # #             writer = pd.ExcelWriter(self.excel_file, engine='xlsxwriter')

# # #             for grocery_title, grocery_data in self.groceries_data.items():
# # #                 flattened_data = []

# # #                 general_info = {
# # #                     'grocery_title': grocery_title,
# # #                     'delivery_time': grocery_data.get('delivery_time', 'N/A'),
# # #                     'delivery_fees': grocery_data.get('grocery_details', {}).get('delivery_fees', 'N/A'),
# # #                     'minimum_order': grocery_data.get('grocery_details', {}).get('minimum_order', 'N/A')
# # #                 }

# # #                 categories = grocery_data.get('grocery_details', {}).get('categories', [])
# # #                 for category in categories:
# # #                     category_name = category.get('name', 'N/A')
# # #                     for sub_category in category.get('sub_categories', []):
# # #                         sub_category_name = sub_category.get('sub_category_name', 'N/A')
# # #                         for item in sub_category.get('Items', []):
# # #                             item_data = {
# # #                                 **general_info,
# # #                                 'category': category_name,
# # #                                 'sub_category': sub_category_name,
# # #                                 'item_name': item.get('item_name', 'N/A'),
# # #                                 'item_price': item.get('item_price', 'N/A'),
# # #                                 'item_description': item.get('item_description', 'N/A'),
# # #                                 'item_link': item.get('item_link', 'N/A')
# # #                             }
# # #                             flattened_data.append(item_data)

# # #                 if not flattened_data:
# # #                     flattened_data.append(general_info)

# # #                 safe_title = re.sub(r'[\\/*?:\[\]]', '_', grocery_title)[:31]
# # #                 df = pd.DataFrame(flattened_data)
# # #                 df.to_excel(writer, sheet_name=safe_title, index=False)

# # #             writer.close()
# # #             print(f"Data saved to {self.excel_file}")
# # #         except Exception as e:
# # #             print(f"Error saving to Excel: {e}")

# # #     async def process_grocery(self, grocery_info):
# # #         """Process a single grocery using TalabatGroceries class"""
# # #         grocery_title = grocery_info['grocery_title']
# # #         grocery_link = grocery_info['grocery_link']
# # #         delivery_time = grocery_info['delivery_time']

# # #         print(f"Processing grocery: {grocery_title}")

# # #         # Skip if already processed
# # #         if grocery_title in self.groceries_data:
# # #             print(f"Skipping {grocery_title} - already processed")
# # #             return

# # #         # Initialize TalabatGroceries with the grocery link
# # #         talabat_grocery = TalabatGroceries(grocery_link)

# # #         # Extract grocery details using Playwright
# # #         for browser_type in ["chromium", "firefox"]:
# # #             try:
# # #                 async with async_playwright() as p:
# # #                     browser = await p[browser_type].launch(headless=True)
# # #                     page = await browser.new_page()

# # #                     grocery_details = await talabat_grocery.extract_categories(page)

# # #                     # Store data in groceries_data dictionary
# # #                     self.groceries_data[grocery_title] = {
# # #                         'grocery_link': grocery_link,
# # #                         'delivery_time': delivery_time,
# # #                         'grocery_details': grocery_details
# # #                     }

# # #                     # Save progress after each grocery
# # #                     self.save_to_json()
# # #                     print(f"Successfully processed and saved data for {grocery_title}")
# # #                     break  # Exit the loop if successful
# # #             except Exception as e:
# # #                 print(f"Error processing {grocery_title} with {browser_type}: {e}")
# # #                 continue  # Try the next browser type
# # #             finally:
# # #                 await browser.close()

# # #     async def run(self):
# # #         """Main method to run the scraper"""
# # #         print("Starting the scraper")
# # #         retries = 3
# # #         while retries > 0:
# # #             try:
# # #                 print(f"Target URL: {self.target_url}")

# # #                 # Initialize playwright and navigate to target URL
# # #                 async with async_playwright() as p:
# # #                     browser = await p.chromium.launch(headless=True)  # Always use headless mode
# # #                     page = await browser.new_page()

# # #                     # Set longer timeouts and wait for page load
# # #                     page.set_default_timeout(240000)  # 240 seconds

# # #                     # Navigate to the target URL
# # #                     await page.goto(self.target_url, timeout=240000)
# # #                     await page.wait_for_load_state("networkidle", timeout=240000)
# # #                     print("Page loaded successfully")

# # #                     # Wait for grocery vendor elements to load
# # #                     try:
# # #                         await page.wait_for_selector('div[data-testid="one-vendor-container"]', timeout=240000)
# # #                         print("Grocery vendor elements found")
# # #                     except Exception as e:
# # #                         print(f"Error waiting for vendor elements: {e}")
# # #                         print("Attempting to continue anyway...")

# # #                     # Extract grocery information directly from page
# # #                     groceries_info = await self.extract_grocery_info(page)
# # #                     await browser.close()

# # #                     print(f"Found {len(groceries_info)} groceries to process")

# # #                     # Process each grocery sequentially
# # #                     for grocery_info in groceries_info:
# # #                         await self.process_grocery(grocery_info)

# # #                     # Save all data to Excel
# # #                     self.save_to_excel()
# # #                     print("Scraping completed successfully")
# # #                 break  # Exit retry loop if successful
# # #             except Exception as e:
# # #                 print(f"Error in main scraper: {e}")
# # #                 retries -= 1
# # #                 print(f"Retries left: {retries}")
# # #                 await asyncio.sleep(5)
# # #                 if retries == 0:
# # #                     print("Failed to complete the scraping process after multiple attempts.")


# # # # Main execution point - now compatible with notebook environments
# # # async def main():
# # #     # Initialize and run the scraper
# # #     scraper = MainScraper()
# # #     await scraper.run()

# # # # Handle both script and notebook execution
# # # if __name__ == "__main__":
# # #     asyncio.run(main())
# # # else:
# # #     # For notebook/IPython environment, use this method to run
# # #     asyncio.get_event_loop().run_until_complete(main())

